import io
import datetime
from flask import Blueprint, send_file, session, redirect, url_for, request
from database import get_db
from functools import wraps

export_bp = Blueprint('export', __name__, url_prefix='/export')


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated


def make_xlsx(headers, rows, sheet_name='Sheet1'):
    """헤더 + 데이터로 xlsx BytesIO 반환"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 헤더 스타일
    header_fill = PatternFill('solid', fgColor='1428A0')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='D1D1D6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border

    # 데이터 행
    alt_fill = PatternFill('solid', fgColor='EEF1FB')
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── TC 리뷰 목록 Excel Export ──────────────────────────
@export_bp.route('/tc_list')
@login_required
def tc_list_excel():
    db = get_db()
    project_id    = request.args.get('project_id', type=int)
    status_filter = request.args.get('status', '')
    domain_filter = request.args.get('domain', '')

    where  = ["1=1"]
    params = []
    if project_id:
        where.append("t.project_id=?"); params.append(project_id)
    if status_filter:
        where.append("t.overall_status=?"); params.append(status_filter)
    if domain_filter:
        where.append("t.domain=?"); params.append(domain_filter)

    tc_list = db.execute(f'''
        SELECT t.*, p.name as project_name
        FROM tc_review t
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE {" AND ".join(where)}
        ORDER BY t.project_id, t.tc_id
    ''', params).fetchall()

    # 보드별 상태도 포함
    headers = ['프로젝트', 'TC ID', 'Sub System', 'IP', 'SITL-ID',
               '제목', 'Domain', 'ALM TC 상태', '전체상태', '보드별 상태', 'TC URL']
    rows = []
    for tc in tc_list:
        board_statuses = db.execute('''
            SELECT b.name, rs.status, u.name as reviewer
            FROM review_status rs
            JOIN boards b ON rs.board_id = b.id
            LEFT JOIN users u ON rs.reviewer_id = u.id
            WHERE rs.tc_id = ?
            ORDER BY b.name
        ''', (tc['id'],)).fetchall()
        board_str = ' / '.join([f"{r['name']}:{r['status']}" for r in board_statuses])

        rows.append([
            tc['project_name'] or '',
            tc['tc_id'] or '',
            tc['sub_system'] or '',
            tc['ip'] or '',
            tc['sitl_id'] or '',
            tc['title'] or '',
            tc['domain'] or '',
            tc['alm_status'] or '',
            tc['overall_status'] or '',
            board_str,
            tc['tc_url'] or '',
        ])

    db.close()
    buf = make_xlsx(headers, rows, sheet_name='TC 리뷰 목록')
    today = datetime.date.today().strftime('%Y%m%d')
    filename = f'TC_리뷰목록_{today}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 리뷰 통합 현황 Excel Export ────────────────────────
@export_bp.route('/summary')
@login_required
def summary_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    db = get_db()
    projects = db.execute(
        "SELECT * FROM projects WHERE is_active=1 ORDER BY name"
    ).fetchall()
    members = db.execute(
        "SELECT * FROM users WHERE is_active=1 ORDER BY location, name"
    ).fetchall()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 삭제

    thin   = Side(style='thin',   color='D1D1D6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_COLORS = {
        'Reviewed':     'E8F5E9',
        'In Review':    'E3F2FD',
        'BL in Review': 'FFF8E1',
        'Draft':        'F2F2F7',
    }

    for p in projects:
        boards = db.execute(
            "SELECT * FROM boards WHERE project_id=? AND is_active=1 ORDER BY name",
            (p['id'],)
        ).fetchall()
        total_tc = db.execute(
            "SELECT COUNT(*) FROM tc_review WHERE project_id=?", (p['id'],)
        ).fetchone()[0]

        if total_tc == 0:
            continue

        ws = wb.create_sheet(title=p['name'][:31])
        ws.freeze_panes = 'B3'

        # ── 헤더 2행 구성 ──
        # 1행: 프로젝트명 + 보드명(병합)
        # 2행: 팀원 | 보드1(In Review / BL / Reviewed) | 보드2 ...
        blue_fill  = PatternFill('solid', fgColor='1428A0')
        blue2_fill = PatternFill('solid', fgColor='1E3FC2')
        white_font = Font(bold=True, color='FFFFFF', size=10)
        center     = Alignment(horizontal='center', vertical='center')

        # 1행
        ws.cell(1, 1, '팀원').fill = blue_fill
        ws.cell(1, 1).font = white_font
        ws.cell(1, 1).alignment = center
        ws.cell(1, 1).border = border

        col = 2
        for b in boards:
            ws.cell(1, col, f'🖥️ {b["name"]}')
            ws.cell(1, col).fill = blue2_fill
            ws.cell(1, col).font = white_font
            ws.cell(1, col).alignment = center
            ws.cell(1, col).border = border
            if len(boards) > 0:
                ws.merge_cells(start_row=1, start_column=col,
                               end_row=1, end_column=col + 2)
            col += 3

        # 2행
        ws.cell(2, 1, '').fill = blue_fill
        ws.cell(2, 1).border = border
        col = 2
        for b in boards:
            for label in ['🔵 In Review', '🟡 BL', '🟢 Reviewed']:
                c = ws.cell(2, col, label)
                c.fill = blue_fill
                c.font = Font(bold=True, color='FFFFFF', size=9)
                c.alignment = center
                c.border = border
                col += 1

        # 데이터 행
        for row_idx, m in enumerate(members, 3):
            name_cell = ws.cell(row_idx, 1, m['name'])
            name_cell.font = Font(bold=True, size=10)
            name_cell.alignment = Alignment(vertical='center')
            name_cell.border = border

            col = 2
            for b in boards:
                stats = db.execute('''
                    SELECT
                        SUM(CASE WHEN rs.status='In Review'    THEN 1 ELSE 0 END) as in_review,
                        SUM(CASE WHEN rs.status='BL in Review' THEN 1 ELSE 0 END) as bl,
                        SUM(CASE WHEN rs.status='Reviewed'     THEN 1 ELSE 0 END) as reviewed
                    FROM tc_review t
                    LEFT JOIN review_status rs ON rs.tc_id=t.id
                        AND rs.board_id=? AND rs.reviewer_id=?
                    WHERE t.project_id=?
                ''', (b['id'], m['id'], p['id'])).fetchone()

                for val, st in [
                    (stats['in_review'], 'In Review'),
                    (stats['bl'],        'BL in Review'),
                    (stats['reviewed'],  'Reviewed'),
                ]:
                    c = ws.cell(row_idx, col, val or 0)
                    c.alignment = center
                    c.border = border
                    if val:
                        c.fill = PatternFill('solid', fgColor=STATUS_COLORS[st])
                        c.font = Font(bold=True, size=10)
                    col += 1

        # 합계 행
        sum_row = len(members) + 3
        sum_cell = ws.cell(sum_row, 1, '합계')
        sum_cell.fill = PatternFill('solid', fgColor='F2F2F7')
        sum_cell.font = Font(bold=True)
        sum_cell.border = border

        col = 2
        for b in boards:
            for st, field in [('In Review','in_review'),('BL in Review','bl'),('Reviewed','reviewed')]:
                total_val = db.execute(f'''
                    SELECT COUNT(*) FROM review_status rs
                    JOIN tc_review t ON rs.tc_id=t.id
                    WHERE t.project_id=? AND rs.board_id=? AND rs.status=?
                ''', (p['id'], b['id'], st)).fetchone()[0]
                c = ws.cell(sum_row, col, total_val)
                c.fill = PatternFill('solid', fgColor='F2F2F7')
                c.font = Font(bold=True)
                c.alignment = center
                c.border = border
                col += 1

        # 열 너비
        ws.column_dimensions['A'].width = 14
        for i in range(2, col):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = 12

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20

    db.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.date.today().strftime('%Y%m%d')
    filename = f'TC_리뷰통합현황_{today}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
